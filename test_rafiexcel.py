import unittest
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PIL import Image as PILImage
import os
import io
from rafiexcel import RafiExcel

class TestRafiExcel(unittest.TestCase):

    def setUp(self):
        self.rafi_excel = RafiExcel()
        self.workbook = Workbook()
        self.ws = self.workbook.active

    def test_merge_and_center_text(self):
        self.rafi_excel.merge_and_center_text(
            column='A1',
            range_col='A1:C1',
            ws=self.ws,
            title='Test Title'
        )
        self.assertEqual(self.ws['A1'].value, 'Test Title')
        self.assertTrue(self.ws['A1'].alignment.horizontal == 'center')
        self.assertTrue(self.ws['A1'].alignment.vertical == 'center')
        self.assertTrue(any("A1:C1" in str(rng) for rng in self.ws.merged_cells.ranges))

    def test_merge_and_left_text_with_color(self):
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.rafi_excel.merge_and_left_text(
            column='B1',
            range_col='B1:D1',
            ws=self.ws,
            title='Left Aligned Title',
            color=fill
        )
        self.assertEqual(self.ws['B1'].value, 'Left Aligned Title')
        self.assertTrue(self.ws['B1'].alignment.horizontal == 'left')
        self.assertEqual(self.ws['B1'].fill.start_color.rgb, fill.start_color.rgb)
        self.assertTrue(any("B1:D1" in str(rng) for rng in self.ws.merged_cells.ranges))

    def test_get_col_name(self):
        col_name = self.rafi_excel.get_col_name(28, 3)
        self.assertEqual(col_name, 'AB3')

    def test_get_col_index(self):
        col_index = self.rafi_excel.get_col_index(28)
        self.assertEqual(col_index, 'AB')

    def test_insert_gambar(self):
        img = PILImage.new('RGB', (100, 100), color='red')
        img_path = 'test_image.jpg'
        img.save(img_path)

        try:
            self.rafi_excel.insert_gambar_custom(
                ws=self.ws,
                column='E5',
                column_id='E',
                row=5,
                path_file=img_path  # Tambahkan argumen yang diperlukan
            )
            images = self.ws._images
            self.assertTrue(len(images) > 0)
        finally:
            if os.path.exists(img_path):
                os.remove(img_path)

    def test_tilt_text(self):
        self.rafi_excel.tilt_text(
            ws=self.ws,
            text='Tilted Text',
            column='F1',
            textRotation=45
        )
        self.assertEqual(self.ws['F1'].value, 'Tilted Text')
        self.assertTrue(self.ws['F1'].alignment.textRotation == 45)

    def test_text_center_and_color(self):
        fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        self.rafi_excel.text_center_and_color(
            ws=self.ws,
            text='Centered Text',
            column='G1',
            color=fill
        )
        self.assertEqual(self.ws['G1'].value, 'Centered Text')
        self.assertTrue(self.ws['G1'].alignment.horizontal == 'center')
        self.assertEqual(self.ws['G1'].fill.start_color.rgb, fill.start_color.rgb)

    def test_read_excel_column(self):
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.append(['Header1', 'Header2'])
        ws.append(['Row1_Col1', 'Row1_Col2'])
        ws.append(['Row2_Col1', 'Row2_Col2'])

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        column_data = self.rafi_excel.read_excel_column(excel_file, 'Header2')
        self.assertEqual(column_data, ['Row1_Col2', 'Row2_Col2'])

    def test_excel_to_list_of_dicts(self):
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        headers = ['no', 'id', 'nama', 'alamat', 'expire']
        ws.append(headers)
        ws.append([1, '001', 'John Doe', '123 Elm St', '2025-12-31'])
        ws.append([2, '002', 'Jane Smith', '456 Oak St', '2026-01-15'])

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        data_list = self.rafi_excel.excel_to_list_of_dicts(excel_file)
        expected_data = [
            {'no': 1, 'id': '001', 'nama': 'John Doe', 'alamat': '123 Elm St', 'expire': '2025-12-31'},
            {'no': 2, 'id': '002', 'nama': 'Jane Smith', 'alamat': '456 Oak St', 'expire': '2026-01-15'}
        ]
        self.assertEqual(data_list, expected_data)

if __name__ == '__main__':
    unittest.main()
